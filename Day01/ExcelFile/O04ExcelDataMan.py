
from openpyxl import load_workbook

wb = load_workbook("FirstExcel.xlsx")

wb.active = wb['Players']

ws = wb.active

for row in ws.iter_rows(min_row=7, min_col=1, max_row=11):
    for cell in row:
        if cell.value == "Lara":
            cell.value = "brain lara".upper()

wb.save("FirstExcel.xlsx")