
from datetime import datetime
from openpyxl import Workbook
# Workbook is a class

wb = Workbook()

ws = wb.active
ws.title = "MyWorkSheet"

ws["C5"] = "Python Programming"
ws["D5"] = 185475.89
ws["E5"] = datetime.now()

wb.save("FirstExcel.xlsx")
