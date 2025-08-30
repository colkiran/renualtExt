
from openpyxl import  load_workbook
from openpyxl.chart import Reference, BarChart3D
from openpyxl.chart.label import DataLabelList

wb = load_workbook("FirstExcel.xlsx")

wb.create_sheet("Charts Example")

wb.active = wb['Charts Example']

ws = wb.active

data = [
    ('Products', 'Sales'),
    ('Pepsi', 1390),
    ('Coke', 1150),
    ('Sprite', 685),
    ('Mirinda', 430),
    ('Thumbs up', 1599),
    ('Fanta', 850),
    ('Slice', 990),
    ('Tropicana', 1240)
]

for row in data:
    ws.append(row)

dataref = Reference(ws, min_row = 2, min_col=2, max_row=8)
labelref = Reference(ws, min_row=2, min_col=1, max_row=8)

chart = BarChart3D()
chart.add_data(dataref)
chart.set_categories(labelref)
chart.x_axis.title = "Products"
chart.y_axis.title = "Sales in lakhs"
chart.title = "Baverage Sales"
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal= True

ws.add_chart(chart, "E9")
wb.save("FirstExcel.xlsx")


